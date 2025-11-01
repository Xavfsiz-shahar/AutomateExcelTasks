"""
Docker-совместимая версия обработчика Excel (без xlwings)
Использует openpyxl с максимальным сохранением форматирования
"""

import os
import io
import shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from copy import copy

class ExcelProcessor:
    def __init__(self, template_path):
        self.template_path = template_path
    
    def process_weekly_file(self, weekly_file_path, month, year, existing_monthly_data=None):
        """
        Обрабатывает еженедельный файл с суммированием данных
        Docker-совместимая версия
        """
        month_name = self._get_month_name(int(month))
        monthly_filename = f'monthly_{year}_{month_name}.xlsx'
        
        if not os.path.exists(self.template_path):
            raise Exception('Шаблон месячного отчета не найден.')
        
        temp_monthly = 'temp_monthly.xlsx'
        
        try:
            # Подготовка базы
            if existing_monthly_data:
                with open(temp_monthly, 'wb') as f:
                    f.write(existing_monthly_data)
            else:
                # Полная копия шаблона
                shutil.copy(self.template_path, temp_monthly)
            
            # Загружаем оба файла
            monthly_wb = load_workbook(temp_monthly)
            weekly_wb = load_workbook(weekly_file_path, data_only=True)
            
            # Обновляем с суммированием
            rows_updated = self._update_with_summation(
                monthly_wb.active, 
                weekly_wb.active
            )
            
            weekly_wb.close()
            
            # Сохраняем в память
            output = io.BytesIO()
            monthly_wb.save(output)
            monthly_wb.close()
            file_data = output.getvalue()
            output.close()
            
            # Очистка
            if os.path.exists(temp_monthly):
                os.remove(temp_monthly)
            
            return monthly_filename, file_data, rows_updated
            
        except Exception as e:
            if os.path.exists(temp_monthly):
                os.remove(temp_monthly)
            raise Exception(f'Ошибка обработки: {str(e)}')
    
    def _get_month_name(self, month):
        """Получить название месяца"""
        return f"{month:02d}"
    
    def _update_with_summation(self, monthly_sheet, weekly_sheet):
        """
        Обновляет данные с суммированием
        Числа складываются, текст перезаписывается
        """
        rows_updated = 0
        max_row = min(weekly_sheet.max_row, 100)
        max_col = min(weekly_sheet.max_column, 100)
        
        for row_idx in range(8, max_row + 1):
            # Проверяем наличие данных
            has_data = False
            for col_idx in range(1, min(5, max_col)):
                if weekly_sheet.cell(row_idx, col_idx).value:
                    has_data = True
                    break
            
            if not has_data:
                continue
            
            # Обновляем ячейки с суммированием
            for col_idx in range(1, max_col + 1):
                weekly_cell = weekly_sheet.cell(row_idx, col_idx)
                monthly_cell = monthly_sheet.cell(row_idx, col_idx)
                
                # Пропускаем объединенные ячейки
                if isinstance(monthly_cell, MergedCell):
                    continue
                
                weekly_value = weekly_cell.value
                
                # Если пусто - пропускаем
                if weekly_value is None or weekly_value == '':
                    continue
                
                monthly_value = monthly_cell.value
                
                try:
                    # Если оба числа - суммируем
                    if isinstance(weekly_value, (int, float)) and isinstance(monthly_value, (int, float)):
                        monthly_cell.value = monthly_value + weekly_value
                    # Если текущее значение число, а новое тоже число - суммируем
                    elif isinstance(weekly_value, (int, float)) and (monthly_value is None or monthly_value == '' or monthly_value == 0):
                        monthly_cell.value = weekly_value
                    # Иначе перезаписываем
                    else:
                        monthly_cell.value = weekly_value
                except:
                    # В случае ошибки просто перезаписываем
                    monthly_cell.value = weekly_value
            
            rows_updated += 1
        
        return rows_updated
    
    def get_monthly_stats(self, file_data):
        """Получает статистику по месячному отчету"""
        try:
            wb = load_workbook(io.BytesIO(file_data), read_only=True)
            ws = wb.active
            
            data_rows = 0
            for row_idx in range(8, ws.max_row + 1):
                for col_idx in range(1, 3):
                    if ws.cell(row_idx, col_idx).value:
                        data_rows += 1
                        break
            
            wb.close()
            
            return {
                'total_rows': data_rows,
                'total_columns': ws.max_column
            }
        except Exception as e:
            return {
                'total_rows': 0,
                'total_columns': 0,
                'error': str(e)
            }

