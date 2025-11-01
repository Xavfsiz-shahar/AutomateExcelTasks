import os
import io
import shutil
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime

class ExcelProcessor:
    def __init__(self, template_path):
        self.template_path = template_path
    
    def process_weekly_file(self, weekly_file_path, month, year, existing_monthly_data=None):
        """
        Обрабатывает еженедельный файл и обновляет месячный отчет
        Использует xlwings для полного сохранения форматирования
        """
        month_name = self._get_month_name(int(month))
        monthly_filename = f'monthly_{year}_{month_name}.xlsx'
        
        if not os.path.exists(self.template_path):
            raise Exception('Шаблон месячного отчета не найден.')
        
        # Временные файлы
        temp_monthly = 'temp_monthly.xlsx'
        
        try:
            # Подготавливаем базу
            if existing_monthly_data:
                # Сохраняем существующий из БД
                with open(temp_monthly, 'wb') as f:
                    f.write(existing_monthly_data)
            else:
                # Копируем шаблон
                shutil.copy(self.template_path, temp_monthly)
            
            # Открываем через xlwings (полное сохранение форматирования!)
            app = xw.App(visible=False, add_book=False)
            
            try:
                # Открываем месячный и недельный файлы
                monthly_book = app.books.open(os.path.abspath(temp_monthly))
                weekly_book = app.books.open(os.path.abspath(weekly_file_path))
                
                monthly_sheet = monthly_book.sheets[0]
                weekly_sheet = weekly_book.sheets[0]
                
                # Обновляем данные с суммированием
                rows_updated = self._update_with_summation(monthly_sheet, weekly_sheet)
                
                # Сохраняем и закрываем
                weekly_book.close()
                monthly_book.save()
                monthly_book.close()
                
            finally:
                app.quit()
            
            # Читаем файл в память
            with open(temp_monthly, 'rb') as f:
                file_data = f.read()
            
            # Удаляем временный файл
            if os.path.exists(temp_monthly):
                os.remove(temp_monthly)
            
            return monthly_filename, file_data, rows_updated
            
        except Exception as e:
            # Очистка при ошибке
            if os.path.exists(temp_monthly):
                os.remove(temp_monthly)
            raise Exception(f'Ошибка обработки: {str(e)}')
    
    def _get_month_name(self, month):
        """Получить название месяца"""
        return f"{month:02d}"
    
    def _update_with_summation(self, monthly_sheet, weekly_sheet):
        """
        Обновляет данные с суммированием
        Числовые значения СКЛАДЫВАЮТСЯ
        Текстовые значения перезаписываются
        """
        rows_updated = 0
        
        # Обновляем строки начиная с 8-й
        max_row = 30  # Примерно столько регионов
        max_col = 80  # Максимум колонок
        
        for row_idx in range(8, max_row):
            # Проверяем есть ли данные в недельном файле
            has_data = False
            for col_idx in range(1, 5):
                try:
                    value = weekly_sheet.range(row_idx, col_idx).value
                    if value:
                        has_data = True
                        break
                except:
                    pass
            
            if not has_data:
                continue
            
            # Обновляем ячейки в этой строке
            for col_idx in range(1, max_col):
                try:
                    weekly_value = weekly_sheet.range(row_idx, col_idx).value
                    monthly_value = monthly_sheet.range(row_idx, col_idx).value
                    
                    # Если в недельном файле пусто - пропускаем
                    if weekly_value is None or weekly_value == '':
                        continue
                    
                    # Если это число - СУММИРУЕМ
                    if isinstance(weekly_value, (int, float)) and isinstance(monthly_value, (int, float)):
                        monthly_sheet.range(row_idx, col_idx).value = monthly_value + weekly_value
                    # Иначе просто перезаписываем
                    else:
                        monthly_sheet.range(row_idx, col_idx).value = weekly_value
                        
                except:
                    pass
            
            rows_updated += 1
        
        return rows_updated
    
    def get_monthly_stats(self, file_data):
        """
        Получает статистику по месячному отчету
        """
        try:
            wb = load_workbook(io.BytesIO(file_data), read_only=True)
            ws = wb.active
            
            # Считаем строки с данными
            data_rows = 0
            for row_idx in range(8, ws.max_row + 1):
                for col_idx in range(1, 3):
                    if ws.cell(row_idx, col_idx).value:
                        data_rows += 1
                        break
            
            wb.close()
            
            return {
                'total_rows': data_rows,
                'total_columns': ws.max_column
            }
        except Exception as e:
            return {
                'total_rows': 0,
                'total_columns': 0,
                'error': str(e)
            }
